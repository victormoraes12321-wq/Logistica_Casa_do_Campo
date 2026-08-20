# Importador opcional: copia dados da planilha Excel antiga para o banco SQLite do sistema.
# Uso: python tools/importar_dados_excel.py "caminho/arquivo.xlsm"
import os, sys, sqlite3, hashlib
from datetime import datetime
try:
    import openpyxl
except ImportError:
    print('Instale openpyxl para usar este importador: pip install openpyxl')
    sys.exit(1)
BASE_DIR=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH=os.path.join(BASE_DIR,'data','logistica_casa_do_campo.sqlite3')
def now(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
def nd(v):
    if not v: return None
    if isinstance(v, datetime): return v.date().isoformat()
    s=str(v)
    if len(s)>=10: return s[:10]
    return None

def hp(p): return hashlib.sha256(('casa_do_campo_local_v2:'+str(p)).encode()).hexdigest()

def main(path):
    wb=openpyxl.load_workbook(path,data_only=True,keep_vba=True)
    db=sqlite3.connect(DB_PATH); db.row_factory=sqlite3.Row
    def get_client(name, city='', farm='', uf='MG', route=''):
        name=(str(name or 'Cliente não informado')).strip(); farm=str(farm or '')
        r=db.execute('SELECT id FROM clients WHERE lower(name)=lower(?) AND ifnull(farm_name,"")=?',(name,farm)).fetchone()
        if r: return r['id']
        return db.execute('INSERT INTO clients(name,city,farm_name,address,notes,route_name,active,created_at) VALUES(?,?,?,?,?,?,1,?)',(name,city or '',farm,farm,f'Importado da planilha. UF: {uf}',route or '',now())).lastrowid
    # usuarios
    if 'Configurações' in wb.sheetnames:
        ws=wb['Configurações']
        for r in range(21,ws.max_row+1):
            name=ws.cell(r,6).value; username=ws.cell(r,7).value or ws.cell(r,8).value; role=ws.cell(r,10).value or 'Operacao'; active=ws.cell(r,11).value
            if name and username:
                db.execute('INSERT OR IGNORE INTO users(name,username,password_hash,role,active,created_at) VALUES(?,?,?,?,?,?)',(str(name),str(username).lower().replace(' ','_'),hp('1234'),str(role),1 if active in (True,'Sim',1,None) else 0,now()))
    # rotas/cidades
    if 'Cadastro de Rotas e SLA' in wb.sheetnames:
        ws=wb['Cadastro de Rotas e SLA']
        for r in range(6,ws.max_row+1):
            rn=ws.cell(r,8).value; city=ws.cell(r,9).value; uf=ws.cell(r,10).value; order=ws.cell(r,11).value; active=ws.cell(r,12).value; notes=ws.cell(r,13).value
            if rn and city:
                db.execute('INSERT INTO route_cities(route_name,city,uf,delivery_order,active,notes) VALUES(?,?,?,?,?,?)',(rn,city,uf,order or 0,1 if active in (True,'Sim',1,None) else 0,notes))
    # clientes cadastrados explícitos
    if 'Cadastro de Clientes' in wb.sheetnames:
        ws=wb['Cadastro de Clientes']
        for r in range(6,ws.max_row+1):
            name=ws.cell(r,2).value
            if name:
                get_client(name, ws.cell(r,3).value, ws.cell(r,6).value, ws.cell(r,4).value, ws.cell(r,5).value)
    # pedidos
    if 'Base de Pedidos' in wb.sheetnames:
        ws=wb['Base de Pedidos']
        header=[ws.cell(5,c).value for c in range(1,ws.max_column+1)]
        idx={h:i+1 for i,h in enumerate(header) if h}
        status_map={'Pronto para carga':'Pronto para entrega','Carga sugerida':'Pronto para entrega','Em rota':'Saiu para entrega','Entregue':'Entrega concluída','Entrega parcial':'Entrega com problema'}
        for r in range(6,ws.max_row+1):
            num=ws.cell(r,idx.get('Numero_Pedido',2)).value
            if not num: continue
            internal=ws.cell(r,idx.get('ID_Pedido_Interno',1)).value
            nf=ws.cell(r,idx.get('Numero_NF',3)).value
            client=ws.cell(r,idx.get('Cliente',4)).value
            city=ws.cell(r,idx.get('Cidade',5)).value
            uf=ws.cell(r,idx.get('UF',6)).value
            farm=ws.cell(r,idx.get('Fazenda_Local_Entrega',7)).value
            route=ws.cell(r,idx.get('Rota',8)).value
            weight=ws.cell(r,idx.get('Peso_kg',9)).value or 0
            sale=nd(ws.cell(r,idx.get('Data_Venda_Compra',10)).value)
            inv=nd(ws.cell(r,idx.get('Data_Faturamento',11)).value)
            delivered=nd(ws.cell(r,idx.get('Data_Entrega',12)).value)
            st=ws.cell(r,idx.get('Status_Pedido',13)).value or 'Aguardando faturamento'
            st=status_map.get(str(st),str(st))
            notes=ws.cell(r,idx.get('Observacoes',15)).value
            expected=nd(ws.cell(r,idx.get('Data_Limite_SLA',17)).value)
            cid=get_client(client,city,farm,uf,route)
            order_number=str(num)
            if db.execute('SELECT id FROM orders WHERE order_number=?',(order_number,)).fetchone(): continue
            oid=db.execute('INSERT INTO orders(order_number,external_id,client_id,status,urgency,sale_date,expected_delivery_date,invoice_limit_date,total_value,weight_kg,delivery_address,route_name,city,uf,notes,invoice_number,invoiced_at,delivered_at,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(order_number,internal,cid,st,'Normal',sale,expected,expected,0,float(weight or 0),farm,route,city,uf,notes,str(nf) if nf else None,inv,delivered,now(),now())).lastrowid
            db.execute('INSERT INTO order_items(order_id,product_name,quantity,unit,weight_kg,notes) VALUES(?,?,?,?,?,?)',(oid,'Carga/Pedido importado',float(weight or 0),'kg',float(weight or 0),'Item gerado a partir do peso total da planilha.'))
            db.execute('INSERT INTO order_history(order_id,user_id,old_status,new_status,action,notes,created_at) VALUES(?,?,?,?,?,?,?)',(oid,None,None,st,'Pedido importado da planilha',notes or '',now()))
    # cargas
    if 'Cargas Geradas' in wb.sheetnames:
        ws=wb['Cargas Geradas']
        for r in range(6,ws.max_row+1):
            cid=ws.cell(r,1).value
            if not cid: continue
            rn=ws.cell(r,2).value; date=nd(ws.cell(r,3).value); status=ws.cell(r,4).value or 'Planejada'; weight=ws.cell(r,5).value or 0; cap=ws.cell(r,6).value or 11000; notes=ws.cell(r,15).value
            st={'Sugerida':'Planejada','Confirmada':'Planejada','Em rota':'Em rota','Entregue':'Concluída'}.get(str(status),str(status))
            db.execute('INSERT INTO routes(name,date,status,route_name,total_weight,capacity,notes,created_at) VALUES(?,?,?,?,?,?,?,?)',(str(cid),date,st,rn,float(weight or 0),float(cap or 0),notes,now()))
    # logs
    if 'Histórico Logs' in wb.sheetnames:
        ws=wb['Histórico Logs']
        for r in range(5,ws.max_row+1):
            if ws.cell(r,1).value:
                db.execute('INSERT INTO audit_logs(created_at,user_name,action,module,entity,old_value,new_value,notes) VALUES(?,?,?,?,?,?,?,?)',(str(ws.cell(r,1).value),ws.cell(r,2).value,ws.cell(r,4).value,ws.cell(r,5).value,ws.cell(r,7).value,ws.cell(r,10).value,ws.cell(r,11).value,ws.cell(r,12).value))
    db.commit(); db.close(); print('Importação concluída.')
if __name__=='__main__':
    if len(sys.argv)<2: print('Informe o caminho do XLSM.'); sys.exit(1)
    main(sys.argv[1])
